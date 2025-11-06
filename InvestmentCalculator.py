import sys
import requests
from xml.etree import ElementTree as ET
import matplotlib.pyplot as plt
from PyQt6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QLineEdit,
    QPushButton, QComboBox, QMessageBox, QFileDialog, QCheckBox
)
from PyQt6.QtGui import QPalette, QColor
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


class InvestmentApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("💰 Инвестиционный калькулятор")
        self.setGeometry(200, 200, 700, 800)
        self.init_ui()
        self.apply_dark_theme()
        self.set_tooltips()

    def init_ui(self):
        layout = QVBoxLayout()

        # --- Режим ---
        self.mode_label = QLabel("Выбери режим расчёта:")
        layout.addWidget(self.mode_label)
        self.mode = QComboBox()
        self.mode.addItems([
            "1 — Через сколько лет стану миллионером",
            "2 — Сколько денег будет через заданный срок",
            "3 — Какая ставка нужна, чтобы стать миллионером"
        ])
        layout.addWidget(self.mode)

        # --- Ввод ---
        self.capital_label = QLabel("Начальная сумма (₽):")
        layout.addWidget(self.capital_label)
        self.capital_input = QLineEdit()
        layout.addWidget(self.capital_input)

        self.percent_label = QLabel("Процентная ставка (% годовых):")
        layout.addWidget(self.percent_label)
        self.percent_input = QLineEdit()
        layout.addWidget(self.percent_input)

        self.term_label = QLabel("Срок (в годах):")
        layout.addWidget(self.term_label)
        self.term_input = QLineEdit()
        layout.addWidget(self.term_input)

        # --- Инфляция ---
        self.inflation_checkbox = QCheckBox("Учитывать инфляцию (%)")
        layout.addWidget(self.inflation_checkbox)
        self.inflation_input = QLineEdit()
        self.inflation_input.setPlaceholderText("Например, 5")
        layout.addWidget(self.inflation_input)

        # --- Валюта ---
        self.currency_label = QLabel("Показать результат в валюте:")
        layout.addWidget(self.currency_label)
        self.currency_box = QComboBox()
        self.currency_box.addItems(["₽ Российский рубль", "$ Доллар США", "€ Евро", "¥ Юань"])
        self.currency_box.currentIndexChanged.connect(self.update_exchange_rate)
        layout.addWidget(self.currency_box)

        self.currency_rate_label = QLabel("Текущий курс: —")
        layout.addWidget(self.currency_rate_label)

        self.currency_rate_input = QLineEdit()
        self.currency_rate_input.setPlaceholderText("Можно ввести вручную, если нет интернета")
        layout.addWidget(self.currency_rate_input)

        # --- Кнопки ---
        self.calc_button = QPushButton("🚀 Рассчитать")
        self.calc_button.clicked.connect(self.calculate)
        layout.addWidget(self.calc_button)

        self.save_button = QPushButton("💾 Сохранить график (PNG/PDF)")
        self.save_button.clicked.connect(self.save_chart)
        layout.addWidget(self.save_button)

        self.export_button = QPushButton("📊 Экспорт в Excel")
        self.export_button.clicked.connect(self.export_to_excel)
        layout.addWidget(self.export_button)

        # --- Результат ---
        self.result_label = QLabel("")
        layout.addWidget(self.result_label)

        # --- График ---
        self.figure, self.ax = plt.subplots(facecolor="#121212")
        self.canvas = FigureCanvas(self.figure)
        layout.addWidget(self.canvas)

        self.setLayout(layout)
        self.apply_styles()

    def apply_styles(self):
        self.setStyleSheet("""
            QLabel { color: #E0E0E0; font-size: 14px; }
            QLineEdit { background-color: #1E1E1E; color: #00FFAA; border: 1px solid #333;
                        border-radius: 5px; padding: 4px; }
            QPushButton { background-color: #00B894; color: white; font-weight: bold;
                         padding: 6px; border-radius: 6px; }
            QPushButton:hover { background-color: #00FFAA; color: black; }
            QComboBox { background-color: #1E1E1E; color: #00FFAA; border: 1px solid #333;
                        border-radius: 5px; padding: 4px; }
            QCheckBox { color: #E0E0E0; font-size: 13px; }
        """)

    def apply_dark_theme(self):
        palette = QPalette()
        palette.setColor(QPalette.ColorRole.Window, QColor(18, 18, 18))
        palette.setColor(QPalette.ColorRole.WindowText, QColor(224, 224, 224))
        palette.setColor(QPalette.ColorRole.Base, QColor(30, 30, 30))
        palette.setColor(QPalette.ColorRole.Text, QColor(0, 255, 170))
        self.setPalette(palette)

    def set_tooltips(self):
        # Режим расчёта
        self.mode.setToolTip(
            "Выберите режим расчёта капитала: когда станете миллионером, будущее через срок или нужная ставка.")

        # Ввод данных
        self.capital_input.setToolTip("Введите сумму, которую планируете инвестировать (в ₽).")
        self.percent_input.setToolTip("Введите ожидаемую годовую процентную ставку (например, 5%).")
        self.term_input.setToolTip("Введите срок инвестирования в годах (для режимов 2 и 3).")

        # Инфляция
        self.inflation_checkbox.setToolTip("Если включено, расчет будет учитывать инфляцию.")
        self.inflation_input.setToolTip("Введите ожидаемый процент инфляции (например, 5%).")

        # Валюта
        self.currency_box.setToolTip("Выберите валюту для отображения результата. Курс подтягивается автоматически.")
        self.currency_rate_input.setToolTip("Можно ввести курс вручную, если нет интернета.")

        # Кнопки
        self.calc_button.setToolTip("Запустить расчет по выбранному режиму.")
        self.save_button.setToolTip("Сохранить график роста капитала в PNG или PDF.")
        self.export_button.setToolTip("Экспортировать таблицу роста капитала в Excel (.xlsx).")

        # Результат
        self.result_label.setToolTip("Здесь отображается итог расчета: номинальная и реальная сумма.")

    def update_exchange_rate(self):
        """Получает актуальный курс валют с ЦБ РФ или оставляет возможность ручного ввода"""
        try:
            symbols = ["RUB", "USD", "EUR", "CNY"]
            code = symbols[self.currency_box.currentIndex()]
            if code == "RUB":
                self.currency_rate_label.setText("Текущий курс: 1 ₽ = 1 ₽")
                self.currency_rate_input.setText("1")
                return
            url = "https://www.cbr.ru/scripts/XML_daily.asp"
            xml = requests.get(url, timeout=5).content
            tree = ET.fromstring(xml)
            rate = None
            for valute in tree.findall("Valute"):
                if valute.find("CharCode").text == code:
                    rate = float(valute.find("Value").text.replace(",", "."))
                    nominal = int(valute.find("Nominal").text)
                    rate /= nominal
                    break
            if rate is not None:
                self.currency_rate_label.setText(f"Текущий курс: 1 ₽ = {rate:.4f} {code}")
                self.currency_rate_input.setText(str(rate))
            else:
                raise ValueError("Курс не найден")
        except Exception as e:
            QMessageBox.warning(self, "Ошибка",
                                "Не удалось получить курс валют с ЦБ РФ.\nВведите вручную.\n\n" + str(e))
            self.currency_rate_label.setText("Ошибка загрузки курса 😔")

    def calculate(self):
        try:
            capital = float(self.capital_input.text())
            percent = float(self.percent_input.text())
            rate = float(self.currency_rate_input.text() or 1)
            inflation = float(self.inflation_input.text() or 0) if self.inflation_checkbox.isChecked() else 0
            mode = self.mode.currentIndex()

            months = [0]
            nominal = [capital]
            real = [capital]

            if mode == 0:
                term = 0
                while capital < 1_000_000:
                    capital += capital * (percent / 100) / 12
                    term += 1
                    months.append(term)
                    nominal.append(capital)
                    real.append(capital / ((1 + inflation / 100 / 12) ** term))
                years = term // 12
                months_left = term % 12
                self.result_label.setText(
                    f"💎 До миллиона: {years} лет и {months_left} мес.\n"
                    f"Номинально: {(capital * rate):,.2f}\n"
                    f"Реально: {(real[-1] * rate):,.2f}"
                )
            elif mode == 1:
                years = float(self.term_input.text())
                for m in range(1, int(years * 12) + 1):
                    capital += capital * (percent / 100) / 12
                    nominal.append(capital)
                    real.append(capital / ((1 + inflation / 100 / 12) ** m))
                    months.append(m)
                self.result_label.setText(
                    f"💰 Через {years} лет:\nНоминально: {(capital * rate):,.2f}\nРеально: {(real[-1] * rate):,.2f}"
                )
            elif mode == 2:
                years = float(self.term_input.text())
                term = int(years * 12)
                found = False
                for p in range(1, 1000):
                    test = float(self.capital_input.text())
                    for _ in range(term):
                        test += test * (p / 10 / 100) / 12
                        if self.inflation_checkbox.isChecked():
                            test /= (1 + inflation / 100 / 12)
                    if test >= 1_000_000:
                        percent = p / 10
                        found = True
                        break
                if found:
                    self.result_label.setText(f"📈 Нужная ставка: {percent:.1f}% годовых")
                else:
                    QMessageBox.warning(self, "Результат", "Даже при 100% годовых миллион не набирается 😅")
                    return

            # Сохраняем данные для экспорта
            self.months = months
            self.nominal = nominal
            self.real = real

            # --- График ---
            self.ax.clear()
            self.ax.set_facecolor("#121212")
            self.ax.plot(months, [n * rate for n in nominal], color="#00FFAA", linewidth=2, label="Номинально")
            if self.inflation_checkbox.isChecked():
                self.ax.plot(months, [r * rate for r in real], color="#FFA500", linestyle="--", linewidth=2,
                             label="С учётом инфляции")
            self.ax.set_title("📊 Рост капитала", color="#FFFFFF", fontsize=12)
            self.ax.set_xlabel("Месяцы", color="#AAAAAA")
            self.ax.set_ylabel("Сумма (в выбранной валюте)", color="#AAAAAA")
            self.ax.grid(color="#333333")
            self.ax.tick_params(colors="#AAAAAA")
            self.ax.legend(facecolor="#121212", edgecolor="#333333", labelcolor="#FFFFFF")
            self.canvas.draw()
        except ValueError:
            QMessageBox.warning(self, "Ошибка", "Проверь введённые данные!")

    def save_chart(self):
        path, _ = QFileDialog.getSaveFileName(self, "Сохранить график", "", "PNG (*.png);;PDF (*.pdf)")
        if path:
            self.figure.savefig(path, bbox_inches="tight")
            QMessageBox.information(self, "Готово", f"График сохранён:\n{path}")

    def export_to_excel(self):
        """Экспорт таблицы роста капитала в Excel"""
        try:
            if not hasattr(self, "months") or not hasattr(self, "nominal") or not hasattr(self, "real"):
                QMessageBox.warning(self, "Ошибка", "Сначала выполните расчёт!")
                return
            path, _ = QFileDialog.getSaveFileName(self, "Сохранить Excel", "", "Excel (*.xlsx)")
            if not path:
                return
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Рост капитала"
            headers = ["Месяцы", "Номинальная сумма", "Реальная сумма"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header).font = Font(bold=True)
            for i, (m, n, r) in enumerate(zip(self.months, self.nominal, self.real), start=2):
                ws.cell(row=i, column=1, value=m)
                ws.cell(row=i, column=2, value=n)
                ws.cell(row=i, column=3, value=r)
            for col in range(1, 4):
                ws.column_dimensions[get_column_letter(col)].width = 18
            wb.save(path)
            QMessageBox.information(self, "Готово", f"Расчёт экспортирован в Excel:\n{path}")
        except Exception as e:
            QMessageBox.warning(self, "Ошибка", f"Не удалось экспортировать в Excel:\n{e}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = InvestmentApp()
    window.show()
    sys.exit(app.exec())
